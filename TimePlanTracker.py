from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox


class Project:
    def __init__(self, name, start_date, end_date):
        self.name = name
        self.start_date = start_date
        self.end_date = end_date
        self.stages = []

    def add_stage(self, name, start_date, end_date):
        stage = Stage(name, start_date, end_date)
        self.stages.append(stage)

    def update_stage(self, stage_name, new_end_date):
        for stage in self.stages:
            if stage.name == stage_name:
                stage.end_date = new_end_date
                break

    def get_completed_stages(self):
        completed_stages = []
        today = datetime.now()
        for stage in self.stages:
            if stage.end_date <= today:
                completed_stages.append(stage)
        return completed_stages

    def to_excel_row(self):
        return [self.name, self.start_date.strftime("%d/%m/%Y"), self.end_date.strftime("%d/%m/%Y")]

    def __str__(self):
        return f"Proje Adı: {self.name}\nBaşlangıç Tarihi: {self.start_date.strftime('%d/%m/%Y')}\nBitiş Tarihi: {self.end_date.strftime('%d/%m/%Y')}\n"


class Stage:
    def __init__(self, name, start_date, end_date):
        self.name = name
        self.start_date = start_date
        self.end_date = end_date

    def to_excel_row(self):
        return [self.name, self.start_date.strftime("%d/%m/%Y"), self.end_date.strftime("%d/%m/%Y")]

    def __str__(self):
        return f"Aşama Adı: {self.name}\nBaşlangıç Tarihi: {self.start_date.strftime('%d/%m/%Y')}\nBitiş Tarihi: {self.end_date.strftime('%d/%m/%Y')}\n"


class ProjectManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.projects = []
        self.load_projects()

    def load_projects(self):
        try:
            workbook = load_workbook(self.file_path)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                name = row[0]
                start_date = datetime.strptime(row[1], "%d/%m/%Y")
                end_date = datetime.strptime(row[2], "%d/%m/%Y")
                project = Project(name, start_date, end_date)
                self.projects.append(project)

            workbook.close()
        except FileNotFoundError:
            print("Excel dosyası bulunamadı.")

    def save_projects(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Proje Adı", "Başlangıç Tarihi", "Bitiş Tarihi"])

        for project in self.projects:
            worksheet.append(project.to_excel_row())

        workbook.save(self.file_path)
        workbook.close()

    def add_project(self, project):
        self.projects.append(project)
        self.save_projects()

    def list_projects(self):
        project_list = []
        for project in self.projects:
            project_list.append(str(project))
        return project_list

    def add_stage_to_project(self, project_name, stage_name, start_date, end_date):
        for project in self.projects:
            if project.name == project_name:
                project.add_stage(stage_name, start_date, end_date)
                self.save_projects()
                break

    def update_stage_in_project(self, project_name, stage_name, new_end_date):
        for project in self.projects:
            if project.name == project_name:
                project.update_stage(stage_name, new_end_date)
                self.save_projects()
                break

    def check_completed_stages(self):
        completed_stages_list = []
        for project in self.projects:
            completed_stages = project.get_completed_stages()
            if completed_stages:
                completed_stages_list.append(f"Proje Adı: {project.name}\n")
                for stage in completed_stages:
                    completed_stages_list.append(str(stage))
        return completed_stages_list


class GUI:
    def __init__(self, manager):
        self.manager = manager

        self.window = tk.Tk()
        self.window.title("Proje Yönetimi")
        self.window.geometry("400x300")

        self.label_title = tk.Label(
            self.window, text="Proje Yönetimi", font=("Arial", 16, "bold"))
        self.label_title.pack(pady=10)

        self.frame_buttons = tk.Frame(self.window)
        self.frame_buttons.pack()

        self.button_list_projects = tk.Button(
            self.frame_buttons, text="Projeleri Listele", width=15, command=self.list_projects)
        self.button_list_projects.grid(row=0, column=0, padx=10, pady=5)

        self.button_add_project = tk.Button(
            self.frame_buttons, text="Proje Ekle", width=15, command=self.add_project)
        self.button_add_project.grid(row=0, column=1, padx=10, pady=5)

        self.button_add_stage = tk.Button(
            self.frame_buttons, text="Aşama Ekle", width=15, command=self.add_stage)
        self.button_add_stage.grid(row=1, column=0, padx=10, pady=5)

        self.button_update_stage = tk.Button(
            self.frame_buttons, text="Aşama Güncelle", width=15, command=self.update_stage)
        self.button_update_stage.grid(row=1, column=1, padx=10, pady=5)

        self.button_check_completed = tk.Button(
            self.frame_buttons, text="Tamamlanan Aşamalar", width=15, command=self.check_completed)
        self.button_check_completed.grid(
            row=2, column=0, columnspan=2, padx=10, pady=5)

        self.listbox = tk.Listbox(self.window, width=50, height=10)
        self.listbox.pack(pady=10)

    def list_projects(self):
        self.listbox.delete(0, tk.END)
        projects = self.manager.list_projects()
        for project in projects:
            self.listbox.insert(tk.END, project)

    def add_project(self):
        project_window = tk.Toplevel(self.window)
        project_window.title("Proje Ekle")

        label_name = tk.Label(project_window, text="Proje Adı:")
        label_name.pack()
        entry_name = tk.Entry(project_window, width=30)
        entry_name.pack()

        label_start_date = tk.Label(
            project_window, text="Başlangıç Tarihi (gg/aa/yyyy):")
        label_start_date.pack()
        entry_start_date = tk.Entry(project_window, width=30)
        entry_start_date.pack()

        label_end_date = tk.Label(
            project_window, text="Bitiş Tarihi (gg/aa/yyyy):")
        label_end_date.pack()
        entry_end_date = tk.Entry(project_window, width=30)
        entry_end_date.pack()

        def add_project():
            name = entry_name.get()
            start_date = datetime.strptime(entry_start_date.get(), "%d/%m/%Y")
            end_date = datetime.strptime(entry_end_date.get(), "%d/%m/%Y")

            project = Project(name, start_date, end_date)
            self.manager.add_project(project)

            messagebox.showinfo("Proje Ekle", "Proje eklendi.")
            project_window.destroy()

        button_add = tk.Button(project_window, text="Ekle",
                               width=10, command=add_project)
        button_add.pack(pady=10)

    def add_stage(self):
        stage_window = tk.Toplevel(self.window)
        stage_window.title("Aşama Ekle")

        label_project_name = tk.Label(stage_window, text="Proje Adı:")
        label_project_name.pack()
        entry_project_name = tk.Entry(stage_window, width=30)
        entry_project_name.pack()

        label_stage_name = tk.Label(stage_window, text="Aşama Adı:")
        label_stage_name.pack()
        entry_stage_name = tk.Entry(stage_window, width=30)
        entry_stage_name.pack()

        label_start_date = tk.Label(
            stage_window, text="Başlangıç Tarihi (gg/aa/yyyy):")
        label_start_date.pack()
        entry_start_date = tk.Entry(stage_window, width=30)
        entry_start_date.pack()

        label_end_date = tk.Label(
            stage_window, text="Bitiş Tarihi (gg/aa/yyyy):")
        label_end_date.pack()
        entry_end_date = tk.Entry(stage_window, width=30)
        entry_end_date.pack()

        def add_stage():
            project_name = entry_project_name.get()
            stage_name = entry_stage_name.get()
            start_date = datetime.strptime(entry_start_date.get(), "%d/%m/%Y")
            end_date = datetime.strptime(entry_end_date.get(), "%d/%m/%Y")

            self.manager.add_stage_to_project(
                project_name, stage_name, start_date, end_date)

            messagebox.showinfo("Aşama Ekle", "Aşama eklendi.")
            stage_window.destroy()

        button_add = tk.Button(stage_window, text="Ekle",
                               width=10, command=add_stage)
        button_add.pack(pady=10)

    def update_stage(self):
        update_window = tk.Toplevel(self.window)
        update_window.title("Aşama Güncelle")

        label_project_name = tk.Label(update_window, text="Proje Adı:")
        label_project_name.pack()
        entry_project_name = tk.Entry(update_window, width=30)
        entry_project_name.pack()

        label_stage_name = tk.Label(update_window, text="Aşama Adı:")
        label_stage_name.pack()
        entry_stage_name = tk.Entry(update_window, width=30)
        entry_stage_name.pack()

        label_end_date = tk.Label(
            update_window, text="Yeni Bitiş Tarihi (gg/aa/yyyy):")
        label_end_date.pack()
        entry_end_date = tk.Entry(update_window, width=30)
        entry_end_date.pack()

        def update_stage():
            project_name = entry_project_name.get()
            stage_name = entry_stage_name.get()
            new_end_date = datetime.strptime(entry_end_date.get(), "%d/%m/%Y")

            self.manager.update_stage_in_project(
                project_name, stage_name, new_end_date)

            messagebox.showinfo("Aşama Güncelle", "Aşama güncellendi.")
            update_window.destroy()

        button_update = tk.Button(
            update_window, text="Güncelle", width=10, command=update_stage)
        button_update.pack(pady=10)

    def check_completed(self):
        completed_stages_window = tk.Toplevel(self.window)
        completed_stages_window.title("Tamamlanan Aşamalar")

        completed_stages = self.manager.check_completed_stages()
        if completed_stages:
            for stage in completed_stages:
                label_stage = tk.Label(completed_stages_window, text=stage)
                label_stage.pack(pady=5)
        else:
            label_no_stages = tk.Label(
                completed_stages_window, text="Tamamlanmış aşama bulunamadı.")
            label_no_stages.pack(pady=5)


def main():
    file_path = "TimePlan.xlsx"
    manager = ProjectManager(file_path)

    gui = GUI(manager)
    gui.window.mainloop()


if __name__ == "__main__":
    main()
